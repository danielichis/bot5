from openpyxl import load_workbook
from pathlib import Path
import sys
import datetime
import os
import shutil
import re

def isNaN(num):
    return num!= num

def getCurrentPath():   
    config_name = 'myapp.cfg'
    # determine if application is a script file or frozen exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)
    application_path2 = Path(application_path)
    return application_path2.absolute()

currentPathFolder = getCurrentPath()
currentPathParentFolder = Path(getCurrentPath()).parent
currentPathGrandpaFolder = Path(currentPathParentFolder).parent
currentPathGrandpaFolderParent = Path(currentPathGrandpaFolder).parent
logPath = os.path.join(currentPathGrandpaFolder,"log.txt")

def today():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d.%m.%Y")
    return currenteDateStr

def today2():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d%m%Y")
    return currenteDateStr

def today3():
    fullTime = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=-4)
    currenteDateStr = fullTime.strftime("%d.%m.%Y-%H %M %S")
    return currenteDateStr




def xlsxFormatting(x, n):
    z1 = ''.join([today(), '-F'])
    wb1 = load_workbook(x)
    xlsxFormatedFolder = os.path.join(currentPathParentFolder,"Cuentas recaudadoras 2")
    xlsxFormatedFolder = os.path.join(xlsxFormatedFolder,z1)
    try:        
        os.mkdir(xlsxFormatedFolder)
    except Exception as e:
        print('El archivo ya ha sido creado')
    
    match n:
        case 1:
            xlsxFormatedPath = os.path.join(xlsxFormatedFolder,"CUENTA ETV-F.xlsx")   

        case 2:
             xlsxFormatedPath = os.path.join(xlsxFormatedFolder,"CUENTA BANCO-F.xlsx")

    wb1.save(xlsxFormatedPath)
    wb2 = load_workbook(xlsxFormatedPath)
    ws2 = wb2['CAJAS RECAUDADORAS']
    mergeRangesList = []
    for i in ws2.merged_cell_ranges:
        i = str(i)
        j = i.index(':')
        #print(i, ' ', j)
        if i[0]==i[j+1] and i[0]=='D' or i[0]==i[j+1]=='E' and i[0]=='E':
            if int(i[1:j])>=3:
                mergeRangesList.append(i)
    #print(mergeRangesList[:])

    for k in mergeRangesList:
        #print(k)
        ws2.unmerge_cells(k)
        l = k.index(':')
        
        a = k[1:l]
        b = k[l+2:]
        #print(a)
        #print(b)
        a = int(a)
        b = int(b)
        a+=1
        b+=1
        
        for m in range(a, b):
            n = ''.join([k[0], str(m)])
            #print( ws2[f'{n}'])
            #print(ws2[k[:l]])
            ws2[f'{n}'] = ws2[k[:l]].value



    wb2.save(xlsxFormatedPath)
    return xlsxFormatedPath

def report(n, asignacion1, accountNumberStr2, accountNumberStr1):
    global errorList
    errorList = []

    match n:
        case 1:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' fue migrado correctamente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case 2:
            p = asignacion1 + '-' + accountNumberStr2 + '-' + accountNumberStr1 + ' no fue migrado correctamente, revisar manualmente'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)
        case _:
            p = 'Error-ingresó un número incorrecto a la función report'
            errorList.append(p)
            # print(p)
            writeLog('\n', p, logPath)

def writeLog(s,log,rut):
    pathLog=os.path.join(rut, "logs.txt")
    line=s+str(log)
    print(log)
    with open(pathLog, 'a') as file:
        file.write(line)

def fecha_a_dia(fecha):
    dias_de_la_semana = ["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]
    fecha_split = fecha.split(".")
    fecha_formateada = datetime.date(int(fecha_split[2]), int(fecha_split[1]), int(fecha_split[0]))
    dia = dias_de_la_semana[fecha_formateada.weekday()]
    return dia

def copyANDeraseFile(fileName):
    todayPath = os.path.join(currentPathParentFolder,"Cuentas recaudadoras")
    fileFromTodayPathFrom = os.path.join(todayPath,fileName)
    todayPath = os.path.join(todayPath,today())
    if not os.path.exists(todayPath):
        os.mkdir(todayPath)
    fileName = today3() + ' ' + fileName
    fileFromTodayPathTo = os.path.join(todayPath,fileName)
    shutil.copyfile(fileFromTodayPathFrom, fileFromTodayPathTo)
    os.remove(fileFromTodayPathFrom)

def copyANDeraseFile2(fileName):
    todayPath = os.path.join(currentPathParentFolder,"Migraciones")
    fileFromTodayPathFrom = os.path.join(todayPath,fileName)
    todayPath = os.path.join(todayPath,today())
    if not os.path.exists(todayPath):
        os.mkdir(todayPath)
    fileName = today3() + ' ' + fileName
    fileFromTodayPathTo = os.path.join(todayPath,fileName)
    shutil.copyfile(fileFromTodayPathFrom, fileFromTodayPathTo)
    os.remove(fileFromTodayPathFrom)

def copyFile(fileName):
    fileNamePathFrom = os.path.join(currentPathParentFolder, fileName)
    fileNamePathTo = os.path.join(currentPathParentFolder, "Cuentas recaudadoras")
    fileNamePathTo = os.path.join(fileNamePathTo, fileName)
    shutil.copyfile(fileNamePathFrom, fileNamePathTo)

def copyFile2(fileName):
    fileNamePathFrom = os.path.join(currentPathParentFolder, fileName)
    fileNamePathTo = os.path.join(currentPathParentFolder, "Migraciones")
    fileNamePathTo = os.path.join(fileNamePathTo, fileName)
    shutil.copyfile(fileNamePathFrom, fileNamePathTo)

def ndocTOxlsx(asignacionNdocMigrated, rec, xlsx, logPath):
    xlsx = xlsx + '.xlsx'
    xlsxPath = os.path.join(currentPathParentFolder, 'Migraciones', xlsx)
    try:
        wb = load_workbook(xlsxPath)
    except:
        writeLog('\n', f'El archivo {xlsx} no existe en el directorio o está abierto.', logPath)
    try:
        ws = wb[rec]
        for l in asignacionNdocMigrated:
            counter = 0
            for i in range(1, ws.max_row+1):
                asignacion = l[0]
                ndoc = l[1]
                if asignacion == ws.cell(row=i, column=1).value:
                    ws[f'K{i}'] = ndoc
                    counter = 1
                    break      
            if counter == 0:
                writeLog('\n', f'La asignación {asignacion} no existe en la hoja {rec} del archivo {xlsx}.', logPath)   
        wb.save(xlsxPath)
        wb.close()
    except:
        writeLog('\n', f'La hoja {rec} no existe en el archivo {xlsx}.', logPath)

def add0(num):
  num_str = str(num)
  if len(num_str) == 1:
    return '0' + num_str
  return num_str

def findUsedCellColumn(ws, row):
    for i in range(1, ws.max_column+1):
        if ws.cell(row=row, column=i).value == None:
            continue
        else:
            return i
        
    if i == ws.max_column:
        return -1

def asig_ndoc_meanwhile(asignacionNdocMigrated, rec, moneda, tMigracion, ETVflow, xlsx, logPath):
    xlsx = xlsx + '.xlsx'
    xlsxPath = os.path.join(currentPathParentFolder, 'Migraciones', xlsx)
    try:
        wb = load_workbook(xlsxPath)
    except:
        writeLog('\n', f'El archivo {xlsx} no existe en el directorio o está abierto.', logPath)
        return -1

    try:
        ws = wb[rec]
    except:
        writeLog('\n', f'La hoja {rec} no existe en el archivo {xlsx}.', logPath)
        return -1
    max_row = ws.max_row

    for l, i in enumerate(asignacionNdocMigrated):
        asignacion = i[0]
        ndoc = i[1]

        match tMigracion:
            case 1:
                match moneda:
                    case 'MN':
                        ws.cell(row=max_row + l + 1, column=1).value = asignacion
                        ws.cell(row=max_row + l + 1, column=2).value = ndoc
                    
                    case 'ME':
                        ws.cell(row=max_row + l + 1, column=4).value = asignacion
                        ws.cell(row=max_row + l + 1, column=5).value = ndoc

            case 2:
                match ETVflow:
                    case 1:
                        match moneda:
                            case 'MN':
                                ws.cell(row=max_row + l + 1, column=7).value = asignacion
                                ws.cell(row=max_row + l + 1, column=8).value = ndoc
                    
                            case 'ME':
                                ws.cell(row=max_row + l + 1, column=10).value = asignacion
                                ws.cell(row=max_row + l + 1, column=11).value = ndoc

                    case 2:
                        match moneda:
                            case 'MN':
                                ws.cell(row=max_row + l + 1, column=13).value = asignacion
                                ws.cell(row=max_row + l + 1, column=14).value = ndoc
                    
                            case 'ME':
                                ws.cell(row=max_row + l + 1, column=16).value = asignacion
                                ws.cell(row=max_row + l + 1, column=17).value = ndoc

                    case 3:
                        match moneda:
                            case 'MN':
                                ws.cell(row=max_row + l + 1, column=7).value = asignacion
                                ws.cell(row=max_row + l + 1, column=8).value = ndoc
                    
                            case 'ME':
                                ws.cell(row=max_row + l + 1, column=10).value = asignacion
                                ws.cell(row=max_row + l + 1, column=11).value = ndoc

    wb.save(xlsxPath)


def asig_ndocToMigra(meanwhileXlsx, logPath):
    writeLog('\n', f'{today3()} Iniciando proceso de pegado de asignaciones y ndocs', logPath)
    meanwhileXlsx = meanwhileXlsx + '.xlsx'
    configPath = os.path.join(currentPathParentFolder, 'config.xlsx')
    try:
        wb2 = load_workbook(configPath)
    except:
        writeLog('\n', f'El archivo config.xlsx no existe en el directorio o está abierto.', logPath)
    
    try:
        ws2 = wb2['parametrosInicio']
    except:
        writeLog('\n', f'La hoja parametrosInicio no existe en el archivo config.xlsx.', logPath)
        return -1
    
    migraXlsx = ws2['B10'].value
    wb2.close()
    migraXlsx = migraXlsx + '.xlsx'
    meanwhileXlsxPath = os.path.join(currentPathParentFolder, 'Migraciones', meanwhileXlsx)
    migraXlsxPath = os.path.join(currentPathParentFolder, 'Migraciones', migraXlsx)
    try:
        wb = load_workbook(meanwhileXlsxPath)
        wb1 = load_workbook(migraXlsxPath)
    
    except:
        writeLog('\n', f'El archivo {meanwhileXlsx} o {migraXlsx} no existe en el directorio o está abierto.', logPath)
        return -1

    sheetNames = wb.get_sheet_names()

    for sheetName in sheetNames:
        try:
            ws = wb[sheetName]
            ws1 = wb1[sheetName]
        except:
            writeLog('\n', f'La hoja {sheetName} no existe en el archivo {meanwhileXlsx} o {migraXlsx}.', logPath)
            continue

        for i in range(1, ws.max_row+1):
            asignacion =  None
            column = findUsedCellColumn(ws, i)
            if column == -1:
                continue
            asignacion = ws.cell(row=i, column=column).value
            asignacion = str(asignacion)
            n1 = re.findall(r'\/(\d+)\/', asignacion)
            if bool(n1) == False:
                    continue
            n1 = int(n1[0])
            x = re.findall(r'(.*\/)(\d{2}).*', asignacion)
            print(x)
            asignacion = x[0][0]+x[0][1]

            ndoc = ws.cell(row=i, column=column+1).value
            for j in range(ws1.max_row, 0, -1):
                asignacion2 = None
                print(j)
                asignacion2 = ws1.cell(row=j, column=1).value
                asignacion2 = str(asignacion2)
                n2 = re.findall(r'\/(\d+)\/', asignacion2)
                n3 = re.findall(r'(^.*?)(?=\/)', asignacion2)
                if bool(n3) == False:
                    n3 = ''
                else:
                    n3 = n3[0]
                if bool(n2) == False:
                    continue
                n2 = int(n2[0])
                if n1-n2>19 and ' ' not in n3:
                    writeLog('\n', f'La asignación {asignacion} - {ndoc} NO SE ENCONTRÓ en el archivo {migraXlsx} en la hoja {sheetName}.', logPath)
                    break
                y = re.findall(r'(.*\/)(\d{2}).*', asignacion2)
                print(y)
                if bool(y) == False:
                    continue
                asignacion2 = y[0][0]+y[0][1]

                if asignacion == asignacion2:
                    match column:
                        case 1:
                            ws1[f'D{j+4}'] = ndoc
                            pene = ws1[f'D{j+4}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {pene} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+4} para MN.', logPath)
                        case 4:
                            ws1[f'E{j+7}'] = ndoc
                            vagina = ws1[f'E{j+7}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {vagina} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+7} para ME.', logPath)
                        case 7:
                            ws1[f'D{j+4}'] = ndoc
                            pene = ws1[f'D{j+4}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {pene} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+4} para MN.', logPath)
                        case 10:
                            ws1[f'D{j+7}'] = ndoc
                            vagina = ws1[f'E{j+7}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {vagina} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+7} para ME.', logPath)
                        case 13:
                            ws1[f'E{j+4}'] = ndoc
                            pene = ws1[f'D{j+4}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {pene} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+4} para MN.', logPath)
                        case 16:
                            ws1[f'E{j+7}'] = ndoc
                            vagina = ws1[f'E{j+7}'].value
                            writeLog('\n', f'{i} La asignación {asignacion} - {vagina} se agregó al archivo {migraXlsx} en la hoja {sheetName} en la fila {j+7} para ME.', logPath)                        
                    break
                if j == 1:
                    writeLog('\n', f'La asignación {asignacion} no existe en la hoja {sheetName} del archivo {migraXlsx}.', logPath)   
                    break
                
    wb1.save(migraXlsxPath)
    wb.save(meanwhileXlsxPath)
    wb.close()
    wb1.close()
    writeLog('\n', f'{today3()} Proceso de pegado de asignaciones y ndocs finalizado', logPath)




        

    
    









if __name__=='__main__': 
  ndocTOxlsx('CCAJ-LP08/298/22', '13161561565', 'AG. ACHUMANI', 'MIGRACIONES SGV DICIEMBRE 2022 28.12.2022', r'C:\Users\crist\OneDrive - UNIVERSIDAD NACIONAL DE INGENIERIA\Venado\Cris\Bot5\Cuentas recaudadoras\logs.txt')